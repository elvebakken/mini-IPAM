from __future__ import annotations

import os
import base64
import csv
import json
import hashlib
import secrets
import threading
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, Optional, List, Tuple
from datetime import datetime, timezone

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Response, Request, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse

from PIL import Image

from .models import (
    LoginRequest, MeResponse, CreateVlanRequest, PatchVlanRequest,
    CreateAssignmentRequest, PatchAssignmentRequest, PatchSettingsRequest,
    ChangePasswordRequest, ChangeUsernameRequest, CreateUserRequest,
    PasswordStrengthRequest, Vlan, Assignment, User,
    MfaSetupResponse, MfaCompleteSetupRequest, MfaDisableRequest,
    MfaSetupDuringLoginRequest, MfaCompleteSetupAndLoginRequest,
    MfaVerifyExportRequest, PatchUserRequest, AdminChangePasswordRequest,
    AdminRecoverMfaRequest
)
from .storage import ensure_files, ensure_admin_user, load_data, save_data, load_users, save_users
from .auth import (
    COOKIE_NAME, cookie_params, create_session_token,
    verify_password, hash_password, require_user, require_role,
    generate_csrf_token, set_csrf_cookie, require_csrf, require_csrf_and_role,
    validate_password_strength, calculate_password_strength,
    check_password_history, update_password_history, is_password_expired, get_password_expiration_date,
    generate_mfa_secret, generate_mfa_qr_code, format_mfa_secret_for_display,
    verify_mfa_code, is_mfa_required, MFA_ENABLED, MFA_ENFORCE_ALL,
    MFA_REQUIRED_FOR_EXPORT, MFA_VERIFY_BEFORE_EXPORT, get_secret, SERVER_INSTANCE_ID
)
from .rate_limit import (
    get_client_ip, check_rate_limit, record_failed_attempt, record_successful_login
)
from .ipcalc import parse_network, usable_range, gateway_suggestion, ip_in_subnet, is_network_or_broadcast, next_available_ip, random_available_ip
from .audit import append_audit, utcnow_iso, read_audit_logs, cleanup_audit_logs
from .validation import (
    sanitize_hostname, sanitize_notes, sanitize_tags, sanitize_vlan_name,
    sanitize_username, sanitize_device_type, sanitize_reserved_reason,
    validate_uploaded_image
)


def ulid_like(prefix: str) -> str:
    # lightweight unique id: timestamp + random
    import time, secrets
    return f"{prefix}_{int(time.time()*1000)}_{secrets.token_hex(6)}"

DATA_DIR = Path(os.getenv("DATA_DIR", "/data"))
ensure_files(DATA_DIR)

# Track used export tokens to prevent reuse (single-use enforcement)
# Key: SHA256 hash of token, Value: timestamp when used
_used_export_tokens: Dict[str, float] = {}
# Lock to prevent race conditions when checking and marking tokens as used
_export_token_lock = threading.Lock()

# MFA Setup Session Storage
# Key: session_key (string), Value: {secret, username, created_at, expires_at}
_mfa_setup_sessions: Dict[str, Dict] = {}
_mfa_setup_lock = threading.Lock()

def _get_mfa_setup_session_key(user: dict) -> str:
    """Generate unique session key for MFA setup."""
    import time
    return f"{user['u']}_{int(time.time()*1000)}_{secrets.token_hex(8)}"

def _get_user_setup_session(username: str) -> Optional[str]:
    """Get active setup session key for user."""
    import time
    now = time.time()
    with _mfa_setup_lock:
        for key, session in _mfa_setup_sessions.items():
            if session['username'] == username and session['expires_at'] > now:
                return key
    return None

def _cleanup_expired_mfa_sessions():
    """Remove expired MFA setup sessions."""
    import time
    now = time.time()
    with _mfa_setup_lock:
        expired_keys = [k for k, v in _mfa_setup_sessions.items() if v['expires_at'] < now]
        for key in expired_keys:
            _mfa_setup_sessions.pop(key, None)

def _hash_token(token: str) -> str:
    """Generate a SHA256 hash of the token for tracking."""
    return hashlib.sha256(token.encode('utf-8')).hexdigest()

def _verify_and_mark_export_token(export_token: str, username: str, request: Request) -> Dict[str, Any]:
    """
    Verify export token and mark it as used to prevent reuse.
    Returns token data if valid, raises HTTPException if invalid or already used.
    
    Uses a lock to ensure atomic check-and-mark operation, preventing race conditions
    where multiple requests could reuse the same token simultaneously.
    """
    from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
    import time
    
    token_hash = _hash_token(export_token)
    
    # Use lock to make check-and-mark operation atomic
    with _export_token_lock:
        # Check if token has already been used (atomic check)
        if token_hash in _used_export_tokens:
            # Log attempted reuse
            append_audit(DATA_DIR / "audit.log", {
                "ts": utcnow_iso(),
                "user": username,
                "action": "export_token_reuse_attempt",
                "ip": get_client_ip(request),
                "details": "Attempted to reuse export token"
            })
            raise HTTPException(status_code=403, detail="Export token has already been used")
        
        # Mark token as "pending verification" immediately to prevent concurrent reuse
        # We'll remove it if verification fails
        _used_export_tokens[token_hash] = time.time()
    
    # Verify token signature and expiration (outside lock to avoid blocking other requests)
    serializer = URLSafeTimedSerializer(get_secret(), salt="export-mfa")
    try:
        token_data = serializer.loads(export_token, max_age=300)  # 5 minutes
    except (BadSignature, SignatureExpired):
        # Verification failed - remove from used tokens since it was invalid
        with _export_token_lock:
            _used_export_tokens.pop(token_hash, None)
        raise HTTPException(status_code=403, detail="Export token expired or invalid")
    
    # Verify token is for export and matches current user
    if token_data.get("type") != "export" or token_data.get("u") != username:
        # Invalid token - remove from used tokens
        with _export_token_lock:
            _used_export_tokens.pop(token_hash, None)
        raise HTTPException(status_code=403, detail="Invalid export token")
    
    # Verify server instance ID
    if token_data.get("i") != SERVER_INSTANCE_ID:
        # Invalid instance ID - remove from used tokens
        with _export_token_lock:
            _used_export_tokens.pop(token_hash, None)
        raise HTTPException(status_code=403, detail="Export token expired")
    
    # Token is valid and has been marked as used - update timestamp
    with _export_token_lock:
        _used_export_tokens[token_hash] = time.time()
    
    # Log token usage
    append_audit(DATA_DIR / "audit.log", {
        "ts": utcnow_iso(),
        "user": username,
        "action": "export_token_used",
        "ip": get_client_ip(request),
        "details": f"Export token used for export operation"
    })
    
    # Clean up old tokens (older than 10 minutes) to prevent memory growth
    with _export_token_lock:
        current_time = time.time()
        expired_hashes = [h for h, ts in _used_export_tokens.items() if current_time - ts > 600]
        for h in expired_hashes:
            del _used_export_tokens[h]
    
    return token_data

app = FastAPI(title="Mini-IPAM", version="0.1.0")


@app.middleware("http")
async def enforce_https(request: Request, call_next):
    """
    Enforce HTTPS in production by redirecting HTTP to HTTPS.
    This middleware must run first to catch HTTP requests before processing.
    
    When behind a reverse proxy, the proxy should set X-Forwarded-Proto header.
    """
    # Check if we're in production (COOKIE_SECURE=true indicates HTTPS is expected)
    is_production = os.getenv("COOKIE_SECURE", "false").lower() == "true"
    
    if is_production:
        # Check if request is HTTP (not HTTPS)
        # In production behind a reverse proxy, check X-Forwarded-Proto header first
        # This is the most reliable indicator when behind a proxy
        forwarded_proto = request.headers.get("X-Forwarded-Proto", "").lower()
        url_scheme = request.url.scheme.lower()
        
        # If we detect HTTP (either via scheme or forwarded-proto header), redirect to HTTPS
        is_http = forwarded_proto == "http" or (not forwarded_proto and url_scheme == "http")
        
        if is_http:
            # Get the host from the request (prefer Host header, fallback to URL)
            host = request.headers.get("Host") or request.url.hostname
            if not host:
                host = request.url.hostname
            
            # Build HTTPS URL (preserve path and query string)
            https_url = f"https://{host}{request.url.path}"
            if request.url.query:
                https_url += f"?{request.url.query}"
            
            # Return 301 permanent redirect to HTTPS
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url=https_url, status_code=301)
    
    # Continue with the request if HTTPS or not in production
    response = await call_next(request)
    return response


@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    """Add security headers to all responses."""
    response = await call_next(request)
    
    # Content-Security-Policy: Prevent XSS attacks
    # Allow self-hosted resources, Tailwind CDN, and API calls to same origin
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.tailwindcss.com; "
        "style-src 'self' 'unsafe-inline' https://cdn.tailwindcss.com; "
        "img-src 'self' data:; "
        "font-src 'self' data:; "
        "connect-src 'self'; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self';"
    )
    response.headers["Content-Security-Policy"] = csp
    
    # X-Frame-Options: Prevent clickjacking
    response.headers["X-Frame-Options"] = "DENY"
    
    # X-Content-Type-Options: Prevent MIME type sniffing
    response.headers["X-Content-Type-Options"] = "nosniff"
    
    # Referrer-Policy: Control referrer information
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    
    # Permissions-Policy: Restrict browser features
    # Disable geolocation, microphone, camera, and other sensitive features
    permissions_policy = (
        "geolocation=(), "
        "microphone=(), "
        "camera=(), "
        "payment=(), "
        "usb=(), "
        "magnetometer=(), "
        "gyroscope=(), "
        "accelerometer=()"
    )
    response.headers["Permissions-Policy"] = permissions_policy
    
    # Strict-Transport-Security: Always add HSTS header when HTTPS is enabled (production)
    # This forces browsers to use HTTPS for all future connections
    if os.getenv("COOKIE_SECURE", "false").lower() == "true":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains; preload"
    
    return response


@app.on_event("startup")
def startup_event():
    """Create admin user on startup if needed and validate HTTPS configuration."""
    ensure_admin_user(DATA_DIR)
    
    # Perform audit log maintenance (rotation, compression, cleanup)
    cleanup_audit_logs(DATA_DIR)
    
    # Log MFA configuration
    print("=" * 60)
    print("Mini-IPAM: MFA Configuration")
    print("=" * 60)
    print(f"MFA_ENABLED: {MFA_ENABLED}")
    print(f"MFA_ENFORCE_ALL: {MFA_ENFORCE_ALL}")
    print(f"MFA_REQUIRED_FOR_EXPORT: {MFA_REQUIRED_FOR_EXPORT}")
    print(f"MFA_VERIFY_BEFORE_EXPORT: {MFA_VERIFY_BEFORE_EXPORT}")
    if MFA_ENFORCE_ALL and not MFA_ENABLED:
        print("WARNING: MFA_ENFORCE_ALL is enabled but MFA_ENABLED is False!")
        print("This should not happen - MFA_ENABLED should be auto-enabled.")
    print("=" * 60)
    
    # Validate HTTPS/TLS configuration in production
    cookie_secure = os.getenv("COOKIE_SECURE", "false").lower() == "true"
    environment = os.getenv("ENVIRONMENT", "").lower()
    is_production = environment == "production" or cookie_secure
    
    if is_production:
        # In production, COOKIE_SECURE must be true
        if not cookie_secure:
            import sys
            print("=" * 60, file=sys.stderr)
            print("WARNING: Production mode detected but COOKIE_SECURE is not set to 'true'", file=sys.stderr)
            print("Set COOKIE_SECURE=true in production to enable secure cookies and HTTPS enforcement.", file=sys.stderr)
            print("=" * 60, file=sys.stderr)
        
        # Warn if running directly (should be behind reverse proxy)
        print("=" * 60)
        print("Mini-IPAM: Production Mode")
        print("=" * 60)
        print("HTTPS enforcement: ENABLED")
        print("HSTS header: ENABLED")
        print("Secure cookies: ENABLED" if cookie_secure else "Secure cookies: DISABLED (WARNING)")
        print("=" * 60)
        print("IMPORTANT: Ensure the application is behind a reverse proxy (nginx, Traefik, Caddy)")
        print("that handles TLS termination. The application will redirect HTTP to HTTPS.")
        print("=" * 60)


def audit(user: dict, action: str, entity: str, entity_id: str, vlan_id: Optional[str], before: Any, after: Any):
    entry = {
        "ts": utcnow_iso(),
        "user": user.get("u", "unknown"),
        "action": action,
        "entity": entity,
        "entity_id": entity_id,
        "vlan_id": vlan_id,
    }
    if before is not None:
        entry["before"] = before
    if after is not None:
        entry["after"] = after
    append_audit(DATA_DIR / "audit.log", entry)


@app.get("/api/health")
def health():
    return {"ok": True}


@app.post("/api/auth/password-strength")
def password_strength(payload: PasswordStrengthRequest):
    """Calculate password strength score and provide feedback."""
    if not payload.password:
        return {
            "score": 0,
            "level": "weak",
            "feedback": ["Enter a password to check strength"]
        }
    return calculate_password_strength(payload.password)


# ---------------- AUTH ----------------

@app.post("/api/auth/login")
def login(payload: LoginRequest, request: Request, response: Response):
    # Get client IP for rate limiting
    client_ip = get_client_ip(request)
    username = payload.username.strip() if payload.username and payload.username.strip() else None
    
    # Check rate limits before attempting authentication
    allowed, error_msg = check_rate_limit(DATA_DIR, client_ip, username)
    if not allowed:
        raise HTTPException(status_code=429, detail=error_msg)
    
    users_file = load_users(DATA_DIR)
    user = next((u for u in users_file.users if u.username == payload.username), None)
    
    # Record failed attempt if user doesn't exist or is disabled
    if not user or user.disabled:
        record_failed_attempt(DATA_DIR, client_ip, username)
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Record failed attempt if password is incorrect
    if not verify_password(payload.password, user.password_bcrypt):
        record_failed_attempt(DATA_DIR, client_ip, username)
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Check if MFA is required
    mfa_required = is_mfa_required(user.mfa_enabled if hasattr(user, 'mfa_enabled') else False)
    user_mfa_secret = getattr(user, 'mfa_secret', None)
    
    if mfa_required:
        # Check if user needs to set up MFA first (MFA_ENFORCE_ALL is enabled but user has no secret)
        if not user_mfa_secret and MFA_ENFORCE_ALL:
            # User needs to set up MFA - return setup required response
            return {
                "ok": False,
                "mfa_setup_required": True,
                "message": "MFA setup is required. Please set up two-factor authentication to continue."
            }
        
        # If MFA is required but no code provided, return MFA challenge
        if not payload.mfa_code:
            return {
                "ok": False,
                "mfa_required": True,
                "message": "Two-factor authentication code required"
            }
        
        # Verify MFA code
        if not user_mfa_secret:
            # User should have MFA enabled but secret is missing - this shouldn't happen
            record_failed_attempt(DATA_DIR, client_ip, username)
            raise HTTPException(status_code=401, detail="MFA configuration error")
        
        if not verify_mfa_code(user_mfa_secret, payload.mfa_code):
            record_failed_attempt(DATA_DIR, client_ip, username)
            raise HTTPException(status_code=401, detail="Invalid two-factor authentication code")

    # Check if password is expired
    password_expired, expiration_date = is_password_expired(user.password_changed_at)
    if password_expired:
        user.password_change_required = True
    
    # Record last login time
    user.last_login_at = utcnow_iso()
    
    # Save user data (password expiration, last login)
    save_users(DATA_DIR, users_file)
    
    # Successful login - clear rate limiting state for this username
    record_successful_login(DATA_DIR, client_ip, username)

    token = create_session_token(user.username, user.role)
    response.set_cookie(COOKIE_NAME, token, **cookie_params())
    
    # Set CSRF token cookie
    csrf_token = generate_csrf_token()
    set_csrf_cookie(response, csrf_token)
    
    # Calculate password expiration date for response
    password_expires_at = get_password_expiration_date(user.password_changed_at)
    
    # Get MFA status
    mfa_enabled = getattr(user, 'mfa_enabled', False)
    mfa_secret = getattr(user, 'mfa_secret', None)
    
    # Check if MFA setup is required (MFA_ENFORCE_ALL is enabled but user hasn't set up MFA)
    mfa_setup_required = False
    if MFA_ENABLED and MFA_ENFORCE_ALL:
        if not mfa_enabled or not mfa_secret:
            mfa_setup_required = True
    
    return {
        "ok": True,
        "user": {
            "username": user.username,
            "role": user.role,
            "password_change_required": user.password_change_required or password_expired,
            "password_expires_at": password_expires_at,
            "mfa_enabled": mfa_enabled,
            "mfa_setup_required": mfa_setup_required,
            "mfa_verify_before_export": MFA_ENABLED and MFA_VERIFY_BEFORE_EXPORT,
            "mfa_required_for_export": MFA_ENABLED and MFA_REQUIRED_FOR_EXPORT
        }
    }


@app.post("/api/auth/logout")
def logout(response: Response, _user=Depends(require_csrf)):
    response.delete_cookie(COOKIE_NAME, path="/")
    response.delete_cookie("csrf_token", path="/")
    return {"ok": True}


# ---------------- MFA (2FA) ---------------- 

@app.post("/api/auth/mfa/setup", response_model=MfaSetupResponse)
def mfa_setup(user=Depends(require_csrf), request: Request):
    """Generate MFA secret and QR code for setup."""
    import time
    if not MFA_ENABLED:
        error_msg = "MFA is not enabled"
        if MFA_ENFORCE_ALL:
            error_msg += " (MFA_ENFORCE_ALL is set but MFA_ENABLED is False - please restart the application)"
        raise HTTPException(status_code=403, detail=error_msg)
    
    # Add rate limiting
    client_ip = get_client_ip(request)
    username = user["u"]
    allowed, error_msg = check_rate_limit(DATA_DIR, client_ip, username)
    if not allowed:
        raise HTTPException(status_code=429, detail=error_msg)
    
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.username == user["u"]), None)
    if not db_user or db_user.disabled:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Check for existing active setup session and invalidate it
    existing_session_key = _get_user_setup_session(user['u'])
    if existing_session_key:
        with _mfa_setup_lock:
            _mfa_setup_sessions.pop(existing_session_key, None)
    
    # Generate new secret
    secret = generate_mfa_secret()
    
    # Store secret in server-side session
    session_key = _get_mfa_setup_session_key(user)
    with _mfa_setup_lock:
        _mfa_setup_sessions[session_key] = {
            "secret": secret,
            "username": user["u"],
            "created_at": time.time(),
            "expires_at": time.time() + 600  # 10 minutes
        }
    
    # Cleanup expired sessions
    _cleanup_expired_mfa_sessions()
    
    # Generate QR code
    qr_code_b64 = generate_mfa_qr_code(secret, db_user.username)
    
    # Format secret for manual entry
    manual_entry_key = format_mfa_secret_for_display(secret)
    
    return {
        "session_key": session_key,
        "qr_code": qr_code_b64,
        "manual_entry_key": manual_entry_key
    }


@app.post("/api/auth/mfa/complete-setup")
def mfa_complete_setup(payload: MfaCompleteSetupRequest, user=Depends(require_csrf)):
    """Complete MFA setup: verify code and enable MFA with the stored secret."""
    import time
    if not MFA_ENABLED:
        raise HTTPException(status_code=403, detail="MFA is not enabled")
    
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.username == user["u"]), None)
    if not db_user or db_user.disabled:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Check if user already has MFA enabled
    if getattr(db_user, 'mfa_enabled', False):
        raise HTTPException(status_code=400, detail="MFA setup cannot be completed")
    
    # Get session key from payload
    session_key = payload.session_key
    
    # Look up stored secret from server-side session
    with _mfa_setup_lock:
        if session_key not in _mfa_setup_sessions:
            raise HTTPException(status_code=400, detail="MFA setup session expired or invalid")
        
        setup_session = _mfa_setup_sessions[session_key]
    
    # Validate session
    if setup_session['username'] != user['u']:
        raise HTTPException(status_code=403, detail="Invalid session")
    
    if setup_session['expires_at'] < time.time():
        with _mfa_setup_lock:
            _mfa_setup_sessions.pop(session_key, None)
        raise HTTPException(status_code=400, detail="MFA setup session expired")
    
    # Use stored secret, not provided secret (if any)
    secret = setup_session['secret']
    code = payload.code
    
    # Verify code matches stored secret
    if not verify_mfa_code(secret, code):
        raise HTTPException(status_code=400, detail="Invalid verification code")
    
    # Enable MFA
    db_user.mfa_enabled = True
    db_user.mfa_secret = secret
    
    # Clear setup session
    with _mfa_setup_lock:
        _mfa_setup_sessions.pop(session_key, None)
    
    save_users(DATA_DIR, users_file)
    audit(user, "user.mfa_enable", "user", db_user.id, None, None, {"mfa_enabled": True})
    
    return {"ok": True, "message": "MFA enabled successfully"}


@app.post("/api/auth/mfa/disable")
def mfa_disable(payload: MfaDisableRequest, user=Depends(require_csrf)):
    """Disable MFA for the current user (requires password verification)."""
    if not MFA_ENABLED:
        raise HTTPException(status_code=403, detail="MFA is not enabled")
    
    # Prevent disabling MFA when MFA_ENFORCE_ALL is enabled
    if MFA_ENFORCE_ALL:
        raise HTTPException(status_code=403, detail="MFA cannot be disabled when MFA_ENFORCE_ALL is enabled")
    
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.username == user["u"]), None)
    if not db_user or db_user.disabled:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify password
    if not verify_password(payload.password, db_user.password_bcrypt):
        raise HTTPException(status_code=401, detail="Invalid password")
    
    # Check if MFA is enabled
    if not getattr(db_user, 'mfa_enabled', False):
        raise HTTPException(status_code=400, detail="MFA is not enabled for this user")
    
    # Disable MFA
    db_user.mfa_enabled = False
    db_user.mfa_secret = None
    
    save_users(DATA_DIR, users_file)
    audit(user, "user.mfa_disable", "user", db_user.id, None, {"mfa_enabled": True}, {"mfa_enabled": False})
    
    return {"ok": True, "message": "MFA disabled successfully"}


@app.post("/api/auth/mfa/setup-during-login", response_model=MfaSetupResponse)
def mfa_setup_during_login(payload: MfaSetupDuringLoginRequest, request: Request):
    """
    Generate MFA secret and QR code for setup during login.
    Requires password verification for security.
    This endpoint is used when MFA_ENFORCE_ALL is enabled and user hasn't set up MFA yet.
    """
    import time
    if not MFA_ENABLED:
        raise HTTPException(status_code=403, detail="MFA is not enabled")
    
    if not MFA_ENFORCE_ALL:
        raise HTTPException(status_code=403, detail="This endpoint is only available when MFA_ENFORCE_ALL is enabled")
    
    # Get client IP for rate limiting
    client_ip = get_client_ip(request)
    username = payload.username.strip() if payload.username and payload.username.strip() else None
    
    # Check rate limits
    allowed, error_msg = check_rate_limit(DATA_DIR, client_ip, username)
    if not allowed:
        raise HTTPException(status_code=429, detail=error_msg)
    
    users_file = load_users(DATA_DIR)
    user = next((u for u in users_file.users if u.username == payload.username), None)
    
    # Verify user exists and is not disabled
    if not user or user.disabled:
        record_failed_attempt(DATA_DIR, client_ip, username)
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Verify password
    if not verify_password(payload.password, user.password_bcrypt):
        record_failed_attempt(DATA_DIR, client_ip, username)
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Check if user already has MFA enabled
    if getattr(user, 'mfa_enabled', False) and getattr(user, 'mfa_secret', None):
        raise HTTPException(status_code=400, detail="MFA is already enabled for this user")
    
    # Check for existing active setup session and invalidate it
    existing_session_key = _get_user_setup_session(username)
    if existing_session_key:
        with _mfa_setup_lock:
            _mfa_setup_sessions.pop(existing_session_key, None)
    
    # Generate new secret
    secret = generate_mfa_secret()
    
    # Store secret in server-side session
    session_key = f"{username}_{int(time.time()*1000)}_{secrets.token_hex(8)}"
    with _mfa_setup_lock:
        _mfa_setup_sessions[session_key] = {
            "secret": secret,
            "username": username,
            "created_at": time.time(),
            "expires_at": time.time() + 600  # 10 minutes
        }
    
    # Cleanup expired sessions
    _cleanup_expired_mfa_sessions()
    
    # Generate QR code
    qr_code_b64 = generate_mfa_qr_code(secret, user.username)
    
    # Format secret for manual entry
    manual_entry_key = format_mfa_secret_for_display(secret)
    
    return {
        "session_key": session_key,
        "qr_code": qr_code_b64,
        "manual_entry_key": manual_entry_key
    }


@app.post("/api/auth/mfa/complete-setup-and-login")
def mfa_complete_setup_and_login(payload: MfaCompleteSetupAndLoginRequest, request: Request, response: Response):
    """
    Complete MFA setup during login and create session.
    Verifies password, sets up MFA, and logs the user in.
    This endpoint is used when MFA_ENFORCE_ALL is enabled and user is setting up MFA for the first time.
    """
    import time
    if not MFA_ENABLED:
        raise HTTPException(status_code=403, detail="MFA is not enabled")
    
    if not MFA_ENFORCE_ALL:
        raise HTTPException(status_code=403, detail="This endpoint is only available when MFA_ENFORCE_ALL is enabled")
    
    # Get client IP for rate limiting
    client_ip = get_client_ip(request)
    username = payload.username.strip() if payload.username and payload.username.strip() else None
    
    # Check rate limits
    allowed, error_msg = check_rate_limit(DATA_DIR, client_ip, username)
    if not allowed:
        raise HTTPException(status_code=429, detail=error_msg)
    
    users_file = load_users(DATA_DIR)
    user = next((u for u in users_file.users if u.username == payload.username), None)
    
    # Verify user exists and is not disabled
    if not user or user.disabled:
        record_failed_attempt(DATA_DIR, client_ip, username)
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Verify password
    if not verify_password(payload.password, user.password_bcrypt):
        record_failed_attempt(DATA_DIR, client_ip, username)
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Get session key from payload
    session_key = payload.session_key
    
    # Look up stored secret from server-side session
    with _mfa_setup_lock:
        if session_key not in _mfa_setup_sessions:
            record_failed_attempt(DATA_DIR, client_ip, username)
            raise HTTPException(status_code=400, detail="MFA setup session expired or invalid")
        
        setup_session = _mfa_setup_sessions[session_key]
    
    # Validate session
    if setup_session['username'] != username:
        record_failed_attempt(DATA_DIR, client_ip, username)
        raise HTTPException(status_code=403, detail="Invalid session")
    
    if setup_session['expires_at'] < time.time():
        with _mfa_setup_lock:
            _mfa_setup_sessions.pop(session_key, None)
        record_failed_attempt(DATA_DIR, client_ip, username)
        raise HTTPException(status_code=400, detail="MFA setup session expired")
    
    # Use stored secret, not provided secret (if any)
    secret = setup_session['secret']
    
    # Verify MFA code matches the stored secret
    if not verify_mfa_code(secret, payload.code):
        record_failed_attempt(DATA_DIR, client_ip, username)
        raise HTTPException(status_code=400, detail="Invalid verification code")
    
    # Enable MFA for the user
    user.mfa_enabled = True
    user.mfa_secret = secret
    
    # Clear setup session
    with _mfa_setup_lock:
        _mfa_setup_sessions.pop(session_key, None)
    
    # Check if password is expired
    password_expired, expiration_date = is_password_expired(user.password_changed_at)
    if password_expired:
        user.password_change_required = True
    
    # Record last login time
    user.last_login_at = utcnow_iso()
    
    save_users(DATA_DIR, users_file)
    
    # Create audit entry (using username since we don't have a session user yet)
    audit_entry = {
        "ts": utcnow_iso(),
        "user": username,
        "action": "user.mfa_enable",
        "entity": "user",
        "entity_id": user.id,
        "vlan_id": None,
        "after": {"mfa_enabled": True}
    }
    append_audit(DATA_DIR / "audit.log", audit_entry)
    
    # Successful login - clear rate limiting state
    record_successful_login(DATA_DIR, client_ip, username)
    
    # Create session token
    token = create_session_token(user.username, user.role)
    response.set_cookie(COOKIE_NAME, token, **cookie_params())
    
    # Set CSRF token cookie
    csrf_token = generate_csrf_token()
    set_csrf_cookie(response, csrf_token)
    
    # Calculate password expiration date
    password_expires_at = get_password_expiration_date(user.password_changed_at)
    
    # Get MFA status (MFA was just enabled, so these should reflect that)
    mfa_enabled = getattr(user, 'mfa_enabled', False)
    mfa_secret = getattr(user, 'mfa_secret', None)
    
    # Check if MFA setup is required (should be False since we just enabled it)
    mfa_setup_required = False
    if MFA_ENABLED and MFA_ENFORCE_ALL:
        if not mfa_enabled or not mfa_secret:
            mfa_setup_required = True
    
    return {
        "ok": True,
        "user": {
            "username": user.username,
            "role": user.role,
            "password_change_required": user.password_change_required or password_expired,
            "password_expires_at": password_expires_at,
            "mfa_enabled": mfa_enabled,
            "mfa_setup_required": mfa_setup_required,
            "mfa_verify_before_export": MFA_ENABLED and MFA_VERIFY_BEFORE_EXPORT,
            "mfa_required_for_export": MFA_ENABLED and MFA_REQUIRED_FOR_EXPORT
        }
    }


@app.post("/api/auth/change-password")
def change_password(payload: ChangePasswordRequest, user=Depends(require_csrf)):
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.username == user["u"]), None)
    if not db_user or db_user.disabled:
        raise HTTPException(status_code=404, detail="User not found")
    
    if not verify_password(payload.current_password, db_user.password_bcrypt):
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    
    # Validate password strength (complexity requirements)
    is_valid, error_msg = validate_password_strength(payload.new_password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Check password history
    is_reused, history_error = check_password_history(payload.new_password, db_user.password_history or [])
    if is_reused:
        raise HTTPException(status_code=400, detail=history_error)
    
    # Update password history before changing password
    old_hash = db_user.password_bcrypt
    db_user.password_history = update_password_history(old_hash, db_user.password_history or [])
    
    # Set new password
    db_user.password_bcrypt = hash_password(payload.new_password)
    db_user.password_change_required = False
    db_user.password_changed_at = utcnow_iso()
    
    save_users(DATA_DIR, users_file)
    audit(user, "user.password_change", "user", db_user.id, None, None, {"password_changed": True})
    return {"ok": True}


@app.post("/api/auth/change-username")
def change_username(payload: ChangeUsernameRequest, response: Response, user=Depends(require_csrf)):
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.username == user["u"]), None)
    if not db_user or db_user.disabled:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Sanitize and validate username
    new_username = sanitize_username(payload.new_username)
    if len(new_username) < 3:
        raise HTTPException(status_code=400, detail="Username must be at least 3 characters")
    
    # Check if username already exists
    if any(u.username == new_username and u.id != db_user.id for u in users_file.users):
        raise HTTPException(status_code=409, detail="Username already exists")
    
    old_username = db_user.username
    db_user.username = new_username
    
    save_users(DATA_DIR, users_file)
    audit(user, "user.username_change", "user", db_user.id, None, {"old_username": old_username}, {"new_username": new_username})
    
    # Update session token with new username
    token = create_session_token(new_username, db_user.role)
    response.set_cookie(COOKIE_NAME, token, **cookie_params())
    return {"ok": True, "username": new_username}


@app.get("/api/me", response_model=MeResponse)
def me(response: Response, request: Request, user=Depends(require_user)):
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.username == user["u"]), None)
    password_change_required = db_user.password_change_required if db_user else False
    
    # Check if password is expired
    password_expired, _ = is_password_expired(db_user.password_changed_at if db_user else None)
    if password_expired and db_user:
        password_change_required = True
        db_user.password_change_required = True
        save_users(DATA_DIR, users_file)
    
    # Ensure CSRF token cookie is set (refresh if missing)
    if not request.cookies.get("csrf_token"):
        csrf_token = generate_csrf_token()
        set_csrf_cookie(response, csrf_token)
    
    # Calculate password expiration date
    password_expires_at = get_password_expiration_date(db_user.password_changed_at if db_user else None)
    
    # Get MFA status
    mfa_enabled = getattr(db_user, 'mfa_enabled', False) if db_user else False
    mfa_secret = getattr(db_user, 'mfa_secret', None) if db_user else None
    
    # Check if MFA setup is required (MFA_ENFORCE_ALL is enabled but user hasn't set up MFA)
    mfa_setup_required = False
    if MFA_ENABLED and MFA_ENFORCE_ALL and db_user:
        if not mfa_enabled or not mfa_secret:
            mfa_setup_required = True
    
    return {
        "username": user["u"],
        "role": user["r"],
        "password_change_required": password_change_required,
        "password_expires_at": password_expires_at,
        "mfa_enabled": mfa_enabled,
        "mfa_setup_required": mfa_setup_required,
        "mfa_verify_before_export": MFA_ENABLED and MFA_VERIFY_BEFORE_EXPORT,
        "mfa_required_for_export": MFA_ENABLED and MFA_REQUIRED_FOR_EXPORT
    }


@app.get("/api/users")
def list_users(user=Depends(require_role({"admin"}))):
    """List all users (admin only)."""
    users_file = load_users(DATA_DIR)
    # Return users without password hashes
    return {
        "users": [
            {
                "id": u.id,
                "username": u.username,
                "role": u.role,
                "disabled": u.disabled,
                "created_at": u.created_at,
                "password_change_required": u.password_change_required,
                "mfa_enabled": getattr(u, 'mfa_enabled', False),
                "last_login_at": getattr(u, 'last_login_at', None)
            }
            for u in users_file.users
        ]
    }


@app.post("/api/users")
def create_user(payload: CreateUserRequest, user=Depends(require_csrf_and_role({"admin"}))):
    users_file = load_users(DATA_DIR)
    
    # Check if username already exists
    if any(u.username == payload.username for u in users_file.users):
        raise HTTPException(status_code=409, detail="Username already exists")
    
    # Sanitize and validate username
    username = sanitize_username(payload.username)
    if len(username) < 3:
        raise HTTPException(status_code=400, detail="Username must be at least 3 characters")
    
    # Validate password strength (complexity requirements)
    is_valid, error_msg = validate_password_strength(payload.password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Handle MFA setup if requested
    mfa_secret = None
    mfa_qr_code = None
    mfa_manual_entry_key = None
    mfa_enabled = False
    
    if payload.mfa_enabled:
        if not MFA_ENABLED:
            raise HTTPException(status_code=400, detail="MFA is not enabled on this server")
        
        # Generate MFA secret and QR code
        mfa_secret = generate_mfa_secret()
        mfa_qr_code = generate_mfa_qr_code(mfa_secret, username)
        mfa_manual_entry_key = format_mfa_secret_for_display(mfa_secret)
        mfa_enabled = True
    
    # Create new user
    now = utcnow_iso()
    new_user = User(
        id=ulid_like("user"),
        username=username,
        password_bcrypt=hash_password(payload.password),
        role=payload.role,
        created_at=now,
        disabled=False,
        password_change_required=False,
        password_history=[],
        password_changed_at=now,
        mfa_enabled=mfa_enabled,
        mfa_secret=mfa_secret
    )
    
    users_file.users.append(new_user)
    save_users(DATA_DIR, users_file)
    audit(user, "user.create", "user", new_user.id, None, None, {"username": new_user.username, "role": new_user.role, "mfa_enabled": mfa_enabled})
    
    response_data = {
        "ok": True,
        "user": {
            "id": new_user.id,
            "username": new_user.username,
            "role": new_user.role
        }
    }
    
    # Include MFA setup data if MFA was enabled
    if mfa_enabled:
        import time
        # Store secret in server-side session
        session_key = f"{username}_{int(time.time()*1000)}_{secrets.token_hex(8)}"
        with _mfa_setup_lock:
            _mfa_setup_sessions[session_key] = {
                "secret": mfa_secret,
                "username": username,
                "created_at": time.time(),
                "expires_at": time.time() + 600  # 10 minutes
            }
        
        response_data["mfa_setup"] = {
            "session_key": session_key,
            "qr_code": mfa_qr_code,
            "manual_entry_key": mfa_manual_entry_key
        }
    
    return response_data


@app.patch("/api/users/{user_id}")
def update_user(user_id: str, payload: PatchUserRequest, user=Depends(require_csrf_and_role({"admin"}))):
    """Update user properties (disable/enable, role, username) (admin only)."""
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.id == user_id), None)
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    before = {"disabled": db_user.disabled, "role": db_user.role, "username": db_user.username}
    after = {}
    
    if payload.disabled is not None:
        db_user.disabled = payload.disabled
        after["disabled"] = db_user.disabled
    
    if payload.role is not None:
        # Prevent admin from removing their own admin role
        if db_user.username == user["u"] and payload.role != "admin":
            raise HTTPException(status_code=400, detail="Cannot remove your own admin role")
        db_user.role = payload.role
        after["role"] = db_user.role
    
    if payload.username is not None:
        # Sanitize and validate username
        new_username = sanitize_username(payload.username)
        if len(new_username) < 3:
            raise HTTPException(status_code=400, detail="Username must be at least 3 characters")
        
        # Check if username already exists
        if any(u.username == new_username and u.id != db_user.id for u in users_file.users):
            raise HTTPException(status_code=409, detail="Username already exists")
        
        old_username = db_user.username
        db_user.username = new_username
        after["username"] = new_username
        before["username"] = old_username
    
    save_users(DATA_DIR, users_file)
    audit(user, "user.update", "user", db_user.id, None, before, after)
    return {"ok": True}


@app.delete("/api/users/{user_id}")
def delete_user(user_id: str, user=Depends(require_csrf_and_role({"admin"}))):
    """Delete a user (admin only)."""
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.id == user_id), None)
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prevent admin from deleting themselves
    if db_user.username == user["u"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    before = {"username": db_user.username, "role": db_user.role}
    users_file.users = [u for u in users_file.users if u.id != user_id]
    save_users(DATA_DIR, users_file)
    audit(user, "user.delete", "user", user_id, None, before, None)
    return {"ok": True}


@app.post("/api/users/{user_id}/change-password")
def admin_change_user_password(user_id: str, payload: AdminChangePasswordRequest, user=Depends(require_csrf_and_role({"admin"}))):
    """Change password for another user (admin only)."""
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.id == user_id), None)
    if not db_user or db_user.disabled:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Validate password strength (complexity requirements)
    is_valid, error_msg = validate_password_strength(payload.new_password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Check password history
    is_reused, history_error = check_password_history(payload.new_password, db_user.password_history or [])
    if is_reused:
        raise HTTPException(status_code=400, detail=history_error)
    
    # Update password history before changing password
    old_hash = db_user.password_bcrypt
    db_user.password_history = update_password_history(old_hash, db_user.password_history or [])
    
    # Set new password
    db_user.password_bcrypt = hash_password(payload.new_password)
    db_user.password_change_required = False
    db_user.password_changed_at = utcnow_iso()
    
    save_users(DATA_DIR, users_file)
    audit(user, "user.password_change_admin", "user", db_user.id, None, None, {"password_changed": True})
    return {"ok": True}


@app.post("/api/users/{user_id}/recover-mfa")
def admin_recover_user_mfa(user_id: str, payload: AdminRecoverMfaRequest, user=Depends(require_csrf_and_role({"admin"}))):
    """
    Recover a user's MFA account by disabling MFA and clearing the secret (admin only).
    
    This endpoint allows administrators to help users who are locked out due to:
    - Lost or broken authenticator device
    - MFA state inconsistency
    - Other MFA-related issues
    
    The recovery action:
    - Disables MFA for the user
    - Clears the MFA secret
    - Allows the user to log in without MFA
    - Logs the action in the audit trail
    
    Security: Requires admin role, CSRF protection, and explicit confirmation.
    """
    if not payload.confirm:
        raise HTTPException(status_code=400, detail="Recovery requires explicit confirmation (confirm=true)")
    
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.id == user_id), None)
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prevent admin from recovering their own MFA (they should use the normal disable flow)
    if db_user.username == user["u"]:
        raise HTTPException(status_code=400, detail="Cannot recover your own MFA. Use the MFA disable feature instead.")
    
    # Store before state for audit
    before = {
        "mfa_enabled": db_user.mfa_enabled,
        "mfa_secret_set": db_user.mfa_secret is not None
    }
    
    # Disable MFA and clear secret
    db_user.mfa_enabled = False
    db_user.mfa_secret = None
    
    save_users(DATA_DIR, users_file)
    
    # Audit the recovery action
    after = {
        "mfa_enabled": False,
        "mfa_secret_set": False,
        "recovered_by": user["u"]
    }
    audit(user, "user.mfa_recover", "user", db_user.id, None, before, after)
    
    return {
        "ok": True,
        "message": f"MFA has been disabled for user '{db_user.username}'. They can now log in without MFA."
    }


# ---------------- SETTINGS ----------------

@app.get("/api/settings")
def get_settings(user=Depends(require_user)):
    data = load_data(DATA_DIR)
    return data.settings.model_dump()


@app.patch("/api/settings")
def patch_settings(payload: PatchSettingsRequest, user=Depends(require_csrf_and_role({"admin"}))):
    data = load_data(DATA_DIR)
    before = data.settings.model_dump()

    if payload.type_options is not None:
        data.settings.type_options = payload.type_options
    if payload.gateway_default is not None:
        data.settings.gateway_default = payload.gateway_default  # type: ignore
    if payload.reserved_defaults is not None:
        data.settings.reserved_defaults = payload.reserved_defaults

    save_data(DATA_DIR, data)
    audit(user, "settings.update", "settings", "settings", None, before, data.settings.model_dump())
    return {"ok": True, "settings": data.settings.model_dump()}


# ---------------- VLAN HELPERS ----------------

def derive_vlan(data_settings: dict, vlan: Vlan) -> Dict[str, Any]:
    net = parse_network(vlan.subnet_cidr)
    start, end, total = usable_range(net)
    gw = vlan.gateway_ip
    if not gw and data_settings.get("gateway_default") == "first_usable":
        gw = gateway_suggestion(net)
    return {
        "network": str(net.network_address),
        "broadcast": str(net.broadcast_address),
        "usable_start": start,
        "usable_end": end,
        "total_usable": total,
        "gateway_suggested": gateway_suggestion(net),
        "gateway_ip": gw,
    }

def reserved_set(vlan: Vlan, settings: dict) -> set[str]:
    net = parse_network(vlan.subnet_cidr)
    res = set()
    rd = settings.get("reserved_defaults", {})
    if rd.get("reserve_network", True):
        res.add(str(net.network_address))
    if rd.get("reserve_broadcast", True):
        res.add(str(net.broadcast_address))
    # gateway: if set, reserve it; else reserve suggested when enabled
    if rd.get("reserve_gateway", True):
        gw = vlan.gateway_ip or gateway_suggestion(net)
        if gw:
            res.add(gw)
    for r in vlan.reserved_ips:
        res.add(r.ip)
    return res

def used_set(vlan: Vlan) -> set[str]:
    return {a.ip for a in vlan.assignments if not a.archived}


# ---------------- VLAN CRUD ----------------

@app.get("/api/vlans")
def list_vlans(user=Depends(require_user)):
    data = load_data(DATA_DIR)
    out = []
    for v in data.vlans:
        d = derive_vlan(data.settings.model_dump(), v)
        res = reserved_set(v, data.settings.model_dump())
        used = used_set(v)
        out.append({
            "id": v.id,
            "name": v.name,
            "vlan_id": v.vlan_id,
            "subnet_cidr": v.subnet_cidr,
            "gateway_ip": v.gateway_ip or d.get("gateway_ip"),
            "derived": {
                "total_usable": d["total_usable"],
                "reserved": len(res),
                "used": len(used),
                "free": max(d["total_usable"] - len(used) - max(0, (len(res) - 2)), 0)  # approximate; UI uses counts mainly
            }
        })
    return out


@app.post("/api/vlans")
def create_vlan(payload: CreateVlanRequest, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    
    # Sanitize and validate VLAN name
    sanitized_name = sanitize_vlan_name(payload.name)
    
    # validate cidr
    try:
        net = parse_network(payload.subnet_cidr)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid CIDR: {e}")

    # Forbid /31 and /32 subnets as they have no usable hosts
    if net.prefixlen >= 31:
        raise HTTPException(status_code=400, detail=f"Subnet /{net.prefixlen} has no usable hosts. /31 and /32 subnets are not supported for VLAN creation.")

    now = utcnow_iso()
    v = Vlan(
        id=ulid_like("vlan"),
        name=sanitized_name,
        vlan_id=payload.vlan_id,
        subnet_cidr=str(net),
        gateway_ip=gateway_suggestion(net) if data.settings.gateway_default == "first_usable" else None,
        reserved_ips=[],
        assignments=[],
        created_at=now,
        updated_at=now,
    )
    data.vlans.append(v)
    save_data(DATA_DIR, data)
    audit(user, "vlan.create", "vlan", v.id, v.id, None, {"name": v.name, "subnet_cidr": v.subnet_cidr})
    return v.model_dump()


@app.get("/api/vlans/{vlan_id}")
def get_vlan(vlan_id: str, user=Depends(require_user)):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    d = derive_vlan(data.settings.model_dump(), v)
    res = sorted(list(reserved_set(v, data.settings.model_dump())))
    used = sorted(list(used_set(v)))
    return {
        **v.model_dump(),
        "derived": d,
        "reserved_effective": res,
        "used_effective": used
    }


@app.patch("/api/vlans/{vlan_id}")
def patch_vlan(vlan_id: str, payload: PatchVlanRequest, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    before = v.model_dump()

    if payload.name is not None:
        v.name = sanitize_vlan_name(payload.name)
    if payload.vlan_id is not None:
        v.vlan_id = payload.vlan_id
    if payload.subnet_cidr is not None:
        try:
            net = parse_network(payload.subnet_cidr)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid CIDR: {e}")
        # Forbid /31 and /32 subnets as they have no usable hosts
        if net.prefixlen >= 31:
            raise HTTPException(status_code=400, detail=f"Subnet /{net.prefixlen} has no usable hosts. /31 and /32 subnets are not supported for VLAN creation.")
        v.subnet_cidr = str(net)
        # optional: reset gateway suggestion if not set
        if not v.gateway_ip and data.settings.gateway_default == "first_usable":
            v.gateway_ip = gateway_suggestion(net)
    if payload.gateway_ip is not None:
        if payload.gateway_ip != "" and not ip_in_subnet(payload.gateway_ip, v.subnet_cidr):
            raise HTTPException(status_code=400, detail="Gateway IP must be inside subnet")
        v.gateway_ip = payload.gateway_ip or None

    v.updated_at = utcnow_iso()
    save_data(DATA_DIR, data)
    audit(user, "vlan.update", "vlan", v.id, v.id, {"name": before["name"], "subnet_cidr": before["subnet_cidr"]},
          {"name": v.name, "subnet_cidr": v.subnet_cidr})
    return {"ok": True}


@app.delete("/api/vlans/{vlan_id}")
def delete_vlan(vlan_id: str, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    idx = next((i for i, x in enumerate(data.vlans) if x.id == vlan_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="VLAN not found")
    before = data.vlans[idx].model_dump()
    data.vlans.pop(idx)
    save_data(DATA_DIR, data)
    audit(user, "vlan.delete", "vlan", vlan_id, vlan_id, {"name": before["name"]}, None)
    return {"ok": True}


# ---------------- ASSIGNMENTS ----------------

@app.get("/api/vlans/{vlan_id}/next-available")
def api_next_available(vlan_id: str, user=Depends(require_user)):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")

    res = reserved_set(v, data.settings.model_dump())
    used = used_set(v)
    ip = next_available_ip(v.subnet_cidr, used, res)
    return {"ip": ip}

@app.get("/api/vlans/{vlan_id}/random-available")
def api_random_available(vlan_id: str, user=Depends(require_user)):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")

    res = reserved_set(v, data.settings.model_dump())
    used = used_set(v)
    ip = random_available_ip(v.subnet_cidr, used, res)
    return {"ip": ip}


def normalize_and_validate_assignment(v: Vlan, data_settings: dict, ip: str, assignment_id: Optional[str] = None):
    # cidr-aware check
    if not ip_in_subnet(ip, v.subnet_cidr):
        raise HTTPException(status_code=400, detail="IP is outside VLAN subnet")

    # Block network and broadcast addresses via CIDR logic (not string matching)
    nb_check = is_network_or_broadcast(ip, v.subnet_cidr)
    if nb_check["is_network"]:
        raise HTTPException(status_code=400, detail="IP is the network address for this subnet")
    if nb_check["is_broadcast"]:
        raise HTTPException(status_code=400, detail="IP is the broadcast address for this subnet")

    # Check if subnet has usable hosts (for /31 and /32)
    net = parse_network(v.subnet_cidr)
    if net.prefixlen >= 31:
        raise HTTPException(status_code=400, detail="Subnet has no usable hosts (/31 and /32 subnets cannot have assignments)")

    # prevent duplicates (excluding self on patch)
    for a in v.assignments:
        if a.archived:
            continue
        if a.ip == ip and (assignment_id is None or a.id != assignment_id):
            raise HTTPException(status_code=409, detail="Duplicate IP in this VLAN")

    # reserved block
    res = reserved_set(v, data_settings)
    if ip in res:
        # Allow reserved only if it's explicitly listed as a custom reserved with reason AND not used by assignment
        raise HTTPException(status_code=400, detail="IP is reserved in this VLAN")


@app.post("/api/vlans/{vlan_id}/assignments")
def create_assignment(vlan_id: str, payload: CreateAssignmentRequest, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")

    normalize_and_validate_assignment(v, data.settings.model_dump(), payload.ip)

    # Sanitize all user inputs
    sanitized_hostname = sanitize_hostname(payload.hostname)
    sanitized_type = sanitize_device_type(payload.type)
    sanitized_tags = sanitize_tags(payload.tags)
    sanitized_notes = sanitize_notes(payload.notes)

    now = utcnow_iso()
    a = Assignment(
        id=ulid_like("asgn"),
        ip=payload.ip,
        hostname=sanitized_hostname,
        type=sanitized_type,
        tags=sanitized_tags,
        notes=sanitized_notes,
        icon=payload.icon,
        archived=False,
        created_at=now,
        updated_at=now,
    )
    v.assignments.append(a)
    v.updated_at = now

    save_data(DATA_DIR, data)
    audit(user, "assignment.create", "assignment", a.id, v.id, None, {"ip": a.ip, "hostname": a.hostname, "type": a.type, "tags": a.tags})
    return a.model_dump()


@app.patch("/api/vlans/{vlan_id}/assignments/{assignment_id}")
def patch_assignment(vlan_id: str, assignment_id: str, payload: PatchAssignmentRequest, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    a = next((x for x in v.assignments if x.id == assignment_id), None)
    if not a:
        raise HTTPException(status_code=404, detail="Assignment not found")
    before = {"ip": a.ip, "hostname": a.hostname, "type": a.type, "tags": a.tags, "archived": a.archived}

    if payload.ip is not None:
        normalize_and_validate_assignment(v, data.settings.model_dump(), payload.ip, assignment_id=a.id)
        a.ip = payload.ip
    if payload.hostname is not None:
        a.hostname = sanitize_hostname(payload.hostname)
    if payload.type is not None:
        a.type = sanitize_device_type(payload.type)
    if payload.tags is not None:
        a.tags = sanitize_tags(payload.tags)
    if payload.notes is not None:
        a.notes = sanitize_notes(payload.notes)
    if payload.icon is not None:
        a.icon = payload.icon
    if payload.archived is not None:
        a.archived = payload.archived

    a.updated_at = utcnow_iso()
    v.updated_at = a.updated_at

    save_data(DATA_DIR, data)
    audit(user, "assignment.update", "assignment", a.id, v.id, before, {"ip": a.ip, "hostname": a.hostname, "type": a.type, "tags": a.tags, "archived": a.archived})
    return {"ok": True}


@app.delete("/api/vlans/{vlan_id}/assignments/{assignment_id}")
def delete_assignment(vlan_id: str, assignment_id: str, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    idx = next((i for i, x in enumerate(v.assignments) if x.id == assignment_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="Assignment not found")
    before = v.assignments[idx].model_dump()
    v.assignments.pop(idx)
    v.updated_at = utcnow_iso()
    save_data(DATA_DIR, data)
    audit(user, "assignment.delete", "assignment", assignment_id, v.id, {"ip": before.get("ip")}, None)
    return {"ok": True}


# ---------------- ICON UPLOAD ----------------

ICONS_DIR = DATA_DIR / "icons"
# Ensure icons directory exists for persistent storage
ICONS_DIR.mkdir(parents=True, exist_ok=True)

@app.get("/api/icons/list")
def list_icons(user=Depends(require_user)):
    """List available predefined icons."""
    if not ICONS_DIR.exists():
        return {"icons": []}
    
    icons = []
    # Include both PNG and SVG files
    for pattern in ["*.png", "*.svg"]:
        for file in sorted(ICONS_DIR.glob(pattern)):
            icons.append({
                "name": file.stem,
                "filename": file.name
            })
    return {"icons": icons}

@app.get("/api/icons/{icon_name}")
def get_icon(icon_name: str, user=Depends(require_role({"admin", "readwrite"}))):
    """Load and normalize a predefined icon."""
    # Security: only allow PNG and SVG files, prevent path traversal
    if not (icon_name.endswith(".png") or icon_name.endswith(".svg")) or "/" in icon_name or "\\" in icon_name:
        raise HTTPException(status_code=400, detail="Invalid icon name")
    
    icon_path = ICONS_DIR / icon_name
    if not icon_path.exists() or not icon_path.is_file():
        raise HTTPException(status_code=404, detail="Icon not found")
    
    try:
        # Handle SVG files
        if icon_name.endswith(".svg"):
            import cairosvg
            # Convert SVG to PNG
            png_bytes = cairosvg.svg2png(url=str(icon_path), output_width=256, output_height=256)
            img = Image.open(BytesIO(png_bytes)).convert("RGBA")
        else:
            # Handle PNG files
            with open(icon_path, "rb") as f:
                raw = f.read()
            img = Image.open(BytesIO(raw)).convert("RGBA")
        
        # center-crop to square
        w, h = img.size
        side = min(w, h)
        left = (w - side) // 2
        top = (h - side) // 2
        img = img.crop((left, top, left + side, top + side))
        img = img.resize((256, 256), Image.LANCZOS)

        out = BytesIO()
        img.save(out, format="PNG", optimize=True)
        png_bytes = out.getvalue()
    except ImportError:
        if icon_name.endswith(".svg"):
            raise HTTPException(status_code=500, detail="SVG support requires cairosvg library")
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid image: {e}")

    b64 = base64.b64encode(png_bytes).decode("ascii")
    return {"mime_type": "image/png", "data_base64": b64}

@app.post("/api/icons/normalize")
def normalize_icon(user=Depends(require_csrf_and_role({"admin", "readwrite"})), file: UploadFile = File(...)):
    if not file.content_type or (not file.content_type.startswith("image/") and file.content_type != "image/svg+xml"):
        raise HTTPException(status_code=400, detail="Only images are allowed")

    raw = file.file.read()
    
    # Validate image using magic bytes (more secure than MIME type alone)
    validate_uploaded_image(raw, file.content_type, max_size=2_000_000)

    try:
        # Handle SVG files
        if file.content_type == "image/svg+xml" or (file.filename and file.filename.endswith(".svg")):
            try:
                import cairosvg
                # Convert SVG to PNG using file_obj for better compatibility
                png_bytes = cairosvg.svg2png(file_obj=BytesIO(raw), output_width=256, output_height=256)
                img = Image.open(BytesIO(png_bytes)).convert("RGBA")
            except ImportError:
                raise HTTPException(status_code=500, detail="SVG support requires cairosvg library")
        else:
            # Handle raster images
            img = Image.open(BytesIO(raw)).convert("RGBA")
        
        # center-crop to square
        w, h = img.size
        side = min(w, h)
        left = (w - side) // 2
        top = (h - side) // 2
        img = img.crop((left, top, left + side, top + side))
        img = img.resize((256, 256), Image.LANCZOS)

        out = BytesIO()
        img.save(out, format="PNG", optimize=True)
        png_bytes = out.getvalue()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid image: {e}")

    b64 = base64.b64encode(png_bytes).decode("ascii")
    return {"mime_type": "image/png", "data_base64": b64}

@app.post("/api/icons/upload-multiple")
async def upload_multiple_icons(user=Depends(require_csrf_and_role({"admin"})), files: list[UploadFile] = File(...)):
    """Upload multiple icons to the icons directory (admin only)."""
    if not ICONS_DIR.exists():
        ICONS_DIR.mkdir(parents=True, exist_ok=True)
    
    uploaded = []
    errors = []
    
    for file in files:
        if not file.content_type or (not file.content_type.startswith("image/") and file.content_type != "image/svg+xml"):
            errors.append({"filename": file.filename, "error": "Only images are allowed"})
            continue
        
        try:
            raw = await file.read()
            
            # Validate image using magic bytes (more secure than MIME type alone)
            try:
                validate_uploaded_image(raw, file.content_type, max_size=2_000_000)
            except HTTPException as e:
                errors.append({"filename": file.filename, "error": e.detail})
                continue
            
            # Check if it's an SVG file
            is_svg = file.content_type == "image/svg+xml" or (file.filename and file.filename.endswith(".svg"))
            
            if is_svg:
                # Preserve SVG files as-is
                # Generate safe filename
                safe_name = "".join(c for c in file.filename if c.isalnum() or c in "._- ").strip()
                if not safe_name:
                    safe_name = "icon"
                if not safe_name.endswith(".svg"):
                    safe_name = safe_name.rsplit(".", 1)[0] + ".svg"
                
                # Ensure unique filename
                base_name = safe_name.rsplit(".", 1)[0]
                counter = 1
                final_path = ICONS_DIR / safe_name
                while final_path.exists():
                    final_path = ICONS_DIR / f"{base_name}_{counter}.svg"
                    counter += 1
                
                # Save SVG as-is
                with open(final_path, "wb") as f:
                    f.write(raw)
                uploaded.append({"filename": final_path.name, "name": base_name})
                audit(user, "icon.upload", "icon", final_path.name, None, None, {"filename": final_path.name})
            else:
                # Validate and normalize raster images
                img = Image.open(BytesIO(raw)).convert("RGBA")
                w, h = img.size
                side = min(w, h)
                left = (w - side) // 2
                top = (h - side) // 2
                img = img.crop((left, top, left + side, top + side))
                img = img.resize((256, 256), Image.LANCZOS)
                
                # Generate safe filename
                safe_name = "".join(c for c in file.filename if c.isalnum() or c in "._- ").strip()
                if not safe_name:
                    safe_name = "icon"
                if not safe_name.endswith(".png"):
                    safe_name = safe_name.rsplit(".", 1)[0] + ".png"
                
                # Ensure unique filename
                base_name = safe_name.rsplit(".", 1)[0]
                counter = 1
                final_path = ICONS_DIR / safe_name
                while final_path.exists():
                    final_path = ICONS_DIR / f"{base_name}_{counter}.png"
                    counter += 1
                
                # Save as PNG
                img.save(final_path, format="PNG", optimize=True)
                uploaded.append({"filename": final_path.name, "name": base_name})
                audit(user, "icon.upload", "icon", final_path.name, None, None, {"filename": final_path.name})
        except Exception as e:
            errors.append({"filename": file.filename, "error": str(e)})
    
    return {
        "uploaded": uploaded,
        "errors": errors,
        "total": len(files),
        "success_count": len(uploaded)
    }

@app.delete("/api/icons/{icon_name}")
def delete_icon(icon_name: str, user=Depends(require_csrf_and_role({"admin"}))):
    """Delete an icon from the icons directory (admin only)."""
    # Security: only allow PNG and SVG files, prevent path traversal
    if not (icon_name.endswith(".png") or icon_name.endswith(".svg")) or "/" in icon_name or "\\" in icon_name:
        raise HTTPException(status_code=400, detail="Invalid icon name")
    
    icon_path = ICONS_DIR / icon_name
    if not icon_path.exists() or not icon_path.is_file():
        raise HTTPException(status_code=404, detail="Icon not found")
    
    try:
        icon_path.unlink()
        audit(user, "icon.delete", "icon", icon_name, None, {"filename": icon_name}, None)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete icon: {e}")

@app.get("/icons/{icon_filename}")
async def serve_icon(icon_filename: str):
    """Serve icon files directly. Fallback route in case StaticFiles mount doesn't work."""
    # Security: only allow PNG and SVG files, prevent path traversal
    if not (icon_filename.endswith(".png") or icon_filename.endswith(".svg")) or "/" in icon_filename or "\\" in icon_filename:
        raise HTTPException(status_code=400, detail="Invalid icon name")
    
    icon_path = ICONS_DIR / icon_filename
    if not icon_path.exists() or not icon_path.is_file():
        raise HTTPException(status_code=404, detail="Icon not found")
    
    # Determine content type
    if icon_filename.endswith(".svg"):
        media_type = "image/svg+xml"
    else:
        media_type = "image/png"
    
    return FileResponse(
        path=str(icon_path),
        media_type=media_type,
        filename=icon_filename
    )


# ---------------- EXPORT ----------------

def check_export_mfa_requirements(user: dict) -> Tuple[bool, Optional[str]]:
    """
    Check if user meets MFA requirements for export.
    Returns (is_allowed, error_message)
    """
    if not MFA_ENABLED:
        return True, None
    
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.username == user["u"]), None)
    if not db_user:
        return False, "User not found"
    
    # Check if MFA is required to be enabled for exports
    if MFA_REQUIRED_FOR_EXPORT:
        mfa_enabled = getattr(db_user, 'mfa_enabled', False)
        mfa_secret = getattr(db_user, 'mfa_secret', None)
        if not mfa_enabled or not mfa_secret:
            return False, "MFA must be enabled to export data"
    
    return True, None


@app.post("/api/export/verify-mfa")
def verify_export_mfa(payload: MfaVerifyExportRequest, user=Depends(require_csrf)):
    """
    Verify MFA code before allowing export.
    Returns a temporary token that can be used for export.
    """
    if not MFA_ENABLED or not MFA_VERIFY_BEFORE_EXPORT:
        raise HTTPException(status_code=403, detail="MFA verification for exports is not enabled")
    
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.username == user["u"]), None)
    if not db_user or db_user.disabled:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Check if user has MFA enabled
    mfa_secret = getattr(db_user, 'mfa_secret', None)
    if not mfa_secret:
        raise HTTPException(status_code=400, detail="MFA is not enabled for this user")
    
    # Verify MFA code
    if not verify_mfa_code(mfa_secret, payload.code):
        raise HTTPException(status_code=401, detail="Invalid MFA code")
    
    # Generate a temporary token for export (valid for 5 minutes)
    # Include a nonce to ensure uniqueness and prevent replay attacks
    from itsdangerous import URLSafeTimedSerializer
    serializer = URLSafeTimedSerializer(get_secret(), salt="export-mfa")
    export_token = serializer.dumps({
        "u": user["u"],
        "r": user["r"],
        "i": SERVER_INSTANCE_ID,
        "type": "export",
        "nonce": secrets.token_urlsafe(32)  # Random nonce to prevent replay
    })
    
    return {
        "ok": True,
        "export_token": export_token,
        "expires_in": 300  # 5 minutes
    }


@app.get("/api/export/data")
def export_data(
    request: Request,
    user=Depends(require_role({"admin", "readwrite", "readonly"})),
    export_token: Optional[str] = Query(None)
):
    # Check MFA requirements
    if MFA_ENABLED:
        # If MFA verification is required, check for valid export token
        if MFA_VERIFY_BEFORE_EXPORT:
            if not export_token:
                raise HTTPException(
                    status_code=403,
                    detail="MFA verification required. Please verify your MFA code first."
                )
            
            # Verify export token and mark as used (single-use enforcement)
            _verify_and_mark_export_token(export_token, user["u"], request)
        
        # Check if MFA must be enabled
        if MFA_REQUIRED_FOR_EXPORT:
            is_allowed, error_msg = check_export_mfa_requirements(user)
            if not is_allowed:
                raise HTTPException(status_code=403, detail=error_msg)
    
    path = DATA_DIR / "data.json"
    if not path.exists():
        raise HTTPException(status_code=404, detail="data.json missing")
    return FileResponse(str(path), media_type="application/json", filename="data.json")


@app.get("/api/vlans/{vlan_id}/assignments/export")
def export_assignments(
    vlan_id: str,
    request: Request,
    format: str = Query("csv", regex="^(csv|json|excel)$"),
    search: Optional[str] = None,
    type_filter: Optional[str] = None,
    export_token: Optional[str] = Query(None),
    user=Depends(require_user)
):
    """Export assignments from a VLAN in CSV, JSON, or Excel format with optional filtering."""
    # Check MFA requirements
    if MFA_ENABLED:
        # If MFA verification is required, check for valid export token
        if MFA_VERIFY_BEFORE_EXPORT:
            if not export_token:
                raise HTTPException(
                    status_code=403,
                    detail="MFA verification required. Please verify your MFA code first."
                )
            
            # Verify export token and mark as used (single-use enforcement)
            _verify_and_mark_export_token(export_token, user["u"], request)
        
        # Check if MFA must be enabled
        if MFA_REQUIRED_FOR_EXPORT:
            is_allowed, error_msg = check_export_mfa_requirements(user)
            if not is_allowed:
                raise HTTPException(status_code=403, detail=error_msg)
    
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    
    # Filter assignments
    assignments = [a for a in v.assignments if not a.archived]
    
    # Apply search filter
    if search:
        search_lower = search.lower()
        assignments = [
            a for a in assignments
            if search_lower in a.ip.lower() or
               search_lower in (a.hostname or "").lower() or
               search_lower in a.type.lower() or
               search_lower in (a.notes or "").lower() or
               any(search_lower in tag.lower() for tag in (a.tags or []))
        ]
    
    # Apply type filter
    if type_filter and type_filter != "all":
        assignments = [a for a in assignments if a.type == type_filter]
    
    # Sort by IP
    assignments.sort(key=lambda x: x.ip)
    
    # Prepare data with VLAN context
    export_data = []
    for a in assignments:
        export_data.append({
            "vlan_name": v.name,
            "vlan_id": v.vlan_id,
            "subnet_cidr": v.subnet_cidr,
            "ip": a.ip,
            "hostname": a.hostname,
            "type": a.type,
            "tags": ", ".join(a.tags) if a.tags else "",
            "notes": a.notes,
            "created_at": a.created_at,
            "updated_at": a.updated_at,
        })
    
    if format == "json":
        json_str = json.dumps(export_data, indent=2)
        return Response(
            content=json_str,
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="assignments_{v.name}_{vlan_id}.json"'}
        )
    
    elif format == "csv":
        output = StringIO()
        if export_data:
            writer = csv.DictWriter(output, fieldnames=export_data[0].keys())
            writer.writeheader()
            writer.writerows(export_data)
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="assignments_{v.name}_{vlan_id}.csv"'}
        )
    
    elif format == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Assignments"
            
            # Headers
            if export_data:
                headers = list(export_data[0].keys())
                ws.append(headers)
                
                # Style header row
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Add data rows
                for row_data in export_data:
                    ws.append([row_data.get(h, "") for h in headers])
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="assignments_{v.name}_{vlan_id}.xlsx"'}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel export requires openpyxl library")


@app.post("/api/vlans/{vlan_id}/assignments/import")
def import_assignments(
    vlan_id: str,
    file: UploadFile = File(...),
    user=Depends(require_csrf_and_role({"admin", "readwrite"}))
):
    """Import assignments from CSV, JSON, or Excel file."""
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    
    file_content = file.file.read()
    filename = file.filename or ""
    
    imported = []
    errors = []
    
    try:
        # Determine file type
        if filename.endswith('.json') or file.content_type == 'application/json':
            # JSON import
            try:
                json_data = json.loads(file_content.decode('utf-8'))
                if isinstance(json_data, list):
                    rows = json_data
                else:
                    rows = [json_data]
            except json.JSONDecodeError as e:
                raise HTTPException(status_code=400, detail=f"Invalid JSON: {e}")
        
        elif filename.endswith('.csv') or file.content_type == 'text/csv':
            # CSV import
            csv_content = file_content.decode('utf-8')
            reader = csv.DictReader(StringIO(csv_content))
            rows = list(reader)
        
        elif filename.endswith(('.xlsx', '.xls')) or 'spreadsheet' in (file.content_type or ''):
            # Excel import
            try:
                from openpyxl import load_workbook
                wb = load_workbook(BytesIO(file_content))
                ws = wb.active
                
                # Read headers
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=False):
                    row_dict = {}
                    for i, cell in enumerate(row):
                        if i < len(headers) and headers[i]:
                            row_dict[headers[i]] = cell.value
                    if any(row_dict.values()):  # Skip empty rows
                        rows.append(row_dict)
            except ImportError:
                raise HTTPException(status_code=500, detail="Excel import requires openpyxl library")
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")
        
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Use CSV, JSON, or Excel.")
        
        # Process rows
        now = utcnow_iso()
        for idx, row in enumerate(rows, start=2):  # Start at 2 for Excel compatibility
            try:
                # Extract IP (required)
                ip = str(row.get('ip', row.get('IP', ''))).strip()
                if not ip:
                    errors.append({"row": idx, "error": "IP address is required"})
                    continue
                
                # Validate IP is in subnet
                if not ip_in_subnet(ip, v.subnet_cidr):
                    errors.append({"row": idx, "ip": ip, "error": f"IP {ip} is outside VLAN subnet {v.subnet_cidr}"})
                    continue
                
                # Block network and broadcast addresses via CIDR logic
                nb_check = is_network_or_broadcast(ip, v.subnet_cidr)
                if nb_check["is_network"]:
                    errors.append({"row": idx, "ip": ip, "error": f"IP {ip} is the network address for this subnet"})
                    continue
                if nb_check["is_broadcast"]:
                    errors.append({"row": idx, "ip": ip, "error": f"IP {ip} is the broadcast address for this subnet"})
                    continue
                
                # Check if subnet has usable hosts (for /31 and /32)
                net = parse_network(v.subnet_cidr)
                if net.prefixlen >= 31:
                    errors.append({"row": idx, "ip": ip, "error": f"Subnet has no usable hosts (/31 and /32 subnets cannot have assignments)"})
                    continue
                
                # Check for duplicates
                if any(a.ip == ip and not a.archived for a in v.assignments):
                    errors.append({"row": idx, "ip": ip, "error": f"IP {ip} already exists in this VLAN"})
                    continue
                
                # Check if reserved
                res = reserved_set(v, data.settings.model_dump())
                if ip in res:
                    errors.append({"row": idx, "ip": ip, "error": f"IP {ip} is reserved in this VLAN"})
                    continue
                
                # Extract and sanitize other fields
                hostname_raw = str(row.get('hostname', row.get('Hostname', ''))).strip()
                type_raw = str(row.get('type', row.get('Type', 'server'))).strip() or 'server'
                
                # Handle tags (can be comma-separated string or list)
                tags_raw = row.get('tags', row.get('Tags', ''))
                if isinstance(tags_raw, list):
                    tags_list = [str(t).strip() for t in tags_raw if t]
                else:
                    tags_list = [t.strip() for t in str(tags_raw).split(',') if t.strip()]
                
                notes_raw = str(row.get('notes', row.get('Notes', ''))).strip()
                
                # Sanitize all inputs
                try:
                    hostname = sanitize_hostname(hostname_raw)
                    assignment_type = sanitize_device_type(type_raw)
                    tags = sanitize_tags(tags_list)
                    notes = sanitize_notes(notes_raw)
                except HTTPException as e:
                    errors.append({"row": idx, "error": f"Validation error: {e.detail}"})
                    continue
                
                # Create assignment
                a = Assignment(
                    id=ulid_like("asgn"),
                    ip=ip,
                    hostname=hostname,
                    type=assignment_type,
                    tags=tags,
                    notes=notes,
                    icon=None,
                    archived=False,
                    created_at=now,
                    updated_at=now,
                )
                
                v.assignments.append(a)
                imported.append({"ip": ip, "hostname": hostname})
                
            except Exception as e:
                errors.append({"row": idx, "error": f"Error processing row: {str(e)}"})
        
        if imported:
            v.updated_at = now
            save_data(DATA_DIR, data)
            audit(user, "assignment.import", "assignment", f"bulk_{vlan_id}", v.id, None, {"count": len(imported)})
        
        return {
            "ok": True,
            "imported": len(imported),
            "errors": len(errors),
            "imported_items": imported,
            "error_details": errors
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")


# ---------------- AUDIT LOGS ----------------

@app.get("/api/audit-logs")
def get_audit_logs(
    user=Depends(require_user),
    user_filter: Optional[str] = None,
    action_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 1000
):
    """Get audit logs with optional filtering."""
    audit_path = DATA_DIR / "audit.log"
    entries = read_audit_logs(audit_path, user_filter, action_filter, date_from, date_to, limit)
    return {"entries": entries}


# Serve icons folder - mount BEFORE catch-all route so it takes precedence
# Always mount (directory is created on startup)
app.mount("/icons", StaticFiles(directory=ICONS_DIR), name="icons")

# Serve static UI - mount at the end so API routes take precedence
static_dir = Path(__file__).parent / "static"
app.mount("/", StaticFiles(directory=static_dir, html=True), name="static")
